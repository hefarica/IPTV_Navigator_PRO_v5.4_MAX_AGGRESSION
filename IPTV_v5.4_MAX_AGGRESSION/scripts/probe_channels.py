"""
probe_channels.py — H9 fix del audit E2E.

Sondea los 18,217 canales en hoja `33_CHANNELS_FROM_FRONTEND` y escribe
`last_status`, `content_type`, `ttfb_ms` por canal. Sin esto, el LAB Excel
no puede declararse E2E porque emite canales muertos a la lista M3U8.

USO:
    python probe_channels.py [--xlsm PATH] [--max N] [--concurrency 50]

PRE-REQUISITOS:
    - Excel debe estar CERRADO completamente (lock file)
    - pip install httpx openpyxl

AUTOR: HFRC + agente, 2026-04-26
"""
import sys, time, asyncio, argparse, datetime, shutil
import openpyxl
import httpx

DEFAULT_XLSM = r"C:\Users\HFRC\Downloads\APE_M3U8_LAB_v8_FIXED.xlsm"
SHEET = "33_CHANNELS_FROM_FRONTEND"

# Columnas a escribir (al final). La macro Brain en Excel puede ajustar si los
# headers ya existen — el script detecta y reusa columnas en ese caso.
HEADER_LAST_STATUS = "last_status"
HEADER_CONTENT_TYPE = "content_type"
HEADER_TTFB_MS = "ttfb_ms"
HEADER_PROBE_AT = "probed_at"


async def probe_one(client: httpx.AsyncClient, url: str) -> dict:
    """HTTP HEAD con timeout 10s. Retorna status/content_type/ttfb_ms."""
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return {"status": 0, "content_type": "INVALID_URL", "ttfb_ms": -1}
    t0 = time.time()
    try:
        r = await client.head(url, timeout=10.0, follow_redirects=True)
        return {
            "status": r.status_code,
            "content_type": (r.headers.get("content-type") or "")[:80],
            "ttfb_ms": int((time.time() - t0) * 1000),
        }
    except httpx.TimeoutException:
        return {"status": 0, "content_type": "TIMEOUT", "ttfb_ms": int((time.time() - t0) * 1000)}
    except Exception as e:
        return {"status": 0, "content_type": f"ERR:{type(e).__name__}"[:80], "ttfb_ms": -1}


def find_or_append_columns(ws, header_names: list) -> dict:
    """Detecta columnas con header_names en R1 (o R2/R3 si están allí). Si faltan, las añade al final."""
    # Buscar header en filas 1-3 (algunos workbooks tienen header en row 3)
    header_row = None
    for r in (1, 2, 3):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower()
                    for c in range(1, ws.max_column + 1)]
        # Si hay >= 5 valores no vacíos parecidos a headers, asumir esa fila es header
        if sum(1 for v in row_vals if v) >= 5:
            header_row = r
            break
    if not header_row:
        header_row = 1

    cols = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        if h:
            cols[h] = c

    out = {}
    next_col = ws.max_column + 1
    for hn in header_names:
        existing = cols.get(hn.lower())
        if existing:
            out[hn] = existing
        else:
            ws.cell(row=header_row, column=next_col).value = hn
            out[hn] = next_col
            next_col += 1
    return {"header_row": header_row, "cols": out}


async def probe_all(xlsm_path: str, max_channels: int = None, concurrency: int = 50,
                    url_col_hint: str = "url"):
    print(f"[probe] Cargando {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    if SHEET not in wb.sheetnames:
        print(f"[ERROR] Hoja '{SHEET}' no existe")
        return
    ws = wb[SHEET]
    print(f"[probe] {ws.max_row} filas x {ws.max_column} cols")

    # Detectar/añadir columnas de output
    info = find_or_append_columns(ws, [HEADER_LAST_STATUS, HEADER_CONTENT_TYPE, HEADER_TTFB_MS, HEADER_PROBE_AT])
    header_row = info["header_row"]
    cols_out = info["cols"]
    print(f"[probe] Header en row {header_row}, output cols: {cols_out}")

    # Detectar URL column
    url_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        if h == url_col_hint.lower() or "url" in h:
            url_col = c
            break
    if not url_col:
        # Fallback: buscar primera celda en data row 2 que parezca URL
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row + 1, column=c).value
            if isinstance(v, str) and v.startswith("http"):
                url_col = c
                break
    if not url_col:
        print("[ERROR] No URL column detectada")
        return
    print(f"[probe] URL en col {url_col}")

    # Recolectar URLs a probar
    data_start = header_row + 1
    urls = []
    for r in range(data_start, ws.max_row + 1):
        u = ws.cell(row=r, column=url_col).value
        if u and isinstance(u, str):
            urls.append((r, u))
    if max_channels:
        urls = urls[:max_channels]
    print(f"[probe] {len(urls)} URLs a sondear (concurrency={concurrency})")

    # Probar
    sem = asyncio.Semaphore(concurrency)
    completed = 0
    t_start = time.time()
    probe_at = datetime.datetime.now().isoformat(timespec="seconds")

    async with httpx.AsyncClient(
        headers={
            "User-Agent": "APE-Probe/1.0 (Channel-Validator)",
            "Accept": "*/*",
        },
        limits=httpx.Limits(max_connections=concurrency, max_keepalive_connections=20),
    ) as client:
        async def task(row: int, url: str):
            nonlocal completed
            async with sem:
                res = await probe_one(client, url)
                ws.cell(row=row, column=cols_out[HEADER_LAST_STATUS]).value = res["status"]
                ws.cell(row=row, column=cols_out[HEADER_CONTENT_TYPE]).value = res["content_type"]
                ws.cell(row=row, column=cols_out[HEADER_TTFB_MS]).value = res["ttfb_ms"]
                ws.cell(row=row, column=cols_out[HEADER_PROBE_AT]).value = probe_at
                completed += 1
                if completed % 500 == 0:
                    rate = completed / (time.time() - t_start)
                    eta = (len(urls) - completed) / rate
                    print(f"[probe] {completed}/{len(urls)} ({rate:.0f} req/s, ETA {eta:.0f}s)")

        await asyncio.gather(*[task(r, u) for r, u in urls])

    elapsed = time.time() - t_start
    print(f"[probe] Done {len(urls)} in {elapsed:.1f}s ({len(urls)/elapsed:.0f} req/s)")

    # Backup + save
    bk = xlsm_path + f".bk_probe_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(xlsm_path, bk)
    print(f"[probe] Backup: {bk}")
    out = xlsm_path.replace(".xlsm", "_PROBED.xlsm")
    wb.save(out)
    print(f"[probe] Saved: {out}")
    print("\nNext: cerrar Excel, mover/renombrar el _PROBED.xlsm para reemplazar el original.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsm", default=DEFAULT_XLSM)
    parser.add_argument("--max", type=int, default=None, help="Max canales a probar")
    parser.add_argument("--concurrency", type=int, default=50)
    parser.add_argument("--url-col", default="url", help="Header de columna URL")
    args = parser.parse_args()
    asyncio.run(probe_all(args.xlsm, args.max, args.concurrency, args.url_col))


if __name__ == "__main__":
    main()
